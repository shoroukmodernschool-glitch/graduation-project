from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from datetime import datetime
import os

app = Flask(__name__)
CORS(app)

EXCEL_PATH = "attendance_log.xlsx"


def open_attendance_sheet():
    if not os.path.exists(EXCEL_PATH):
        return None

    workbook = load_workbook(EXCEL_PATH)
    return workbook["Attendance"]


def check_attendance_today(student_id):
    today = datetime.now().strftime("%Y-%m-%d")
    sheet = open_attendance_sheet()

    if sheet is None:
        return "ملف الحضور مش موجود."

    for row in sheet.iter_rows(min_row=2, values_only=True):
        excel_student_id = str(row[0]).strip()
        date = str(row[2]).strip()
        time = str(row[3]).strip()

        if excel_student_id == str(student_id).strip() and date == today:
            return f"حضورك متسجل النهارده الساعة {time}."

    return "حضورك مش متسجل النهارده."


def get_all_attendance_days(student_id):
    sheet = open_attendance_sheet()

    if sheet is None:
        return "ملف الحضور مش موجود."

    attendance_days = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        excel_student_id = str(row[0]).strip()
        date = str(row[2]).strip()
        time = str(row[3]).strip()
        status = str(row[4]).strip()

        if excel_student_id == str(student_id).strip():
            attendance_days.append(f"- {date} | الساعة {time} | {status}")

    if not attendance_days:
        return "مفيش أي حضور متسجل للطالب ده."

    days_text = "\n".join(attendance_days)

    return f"عدد مرات الحضور: {len(attendance_days)}\n\nالأيام:\n{days_text}"


def understand_message(message, student_id):
    message = message.lower().strip()

    # السؤال التاني: كل الأيام وعدد مرات الحضور
    if "how many times" in message:
        return get_all_attendance_days(student_id)

    if "كام مرة" in message or "عدد المرات" in message or "حضرت كام" in message:
        return get_all_attendance_days(student_id)

    # السؤال الأول: النهارده بس
    if "has my attendance been recorded" in message:
        return check_attendance_today(student_id)

    if "am i absent today" in message or "today" in message:
        return check_attendance_today(student_id)

    if "غبت" in message or "غايب" in message or "النهارده" in message or "اليوم" in message:
        return check_attendance_today(student_id)

    return "مش فاهم سؤالك كويس. اسألني عن الحضور مثلًا: هل أنا غبت النهارده؟"


@app.route("/ask", methods=["POST"])
def ask():
    data = request.get_json()

    student_id = data.get("student_id")
    message = data.get("message")

    if not student_id:
        return jsonify({"reply": "student_id مطلوب."}), 400

    if not message:
        return jsonify({"reply": "message مطلوب."}), 400

    reply = understand_message(message, student_id)

    return jsonify({"reply": reply})


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5001, debug=True)