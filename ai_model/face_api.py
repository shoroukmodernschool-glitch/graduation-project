import os
from datetime import datetime
from openpyxl import load_workbook
from flask import Flask, Response, request, jsonify
from flask_cors import CORS
from face_recognition import generate_frames

app = Flask(__name__)
CORS(app)


@app.route("/")
def home():
    return "Face API is running ✅"


@app.route("/check-attendance", methods=["POST"])
def check_attendance():
    data = request.get_json()

    student_id = str(data.get("student_id", "")).strip()

    if not student_id:
        return jsonify({
            "reply": "ممكن تبعت الـ ID الخاص بيك؟"
        })

    file_path = "attendance_log.xlsx"
    today = datetime.now().strftime("%Y-%m-%d")

    if not os.path.exists(file_path):
        return jsonify({
            "reply": "لا، حضورك مش متسجل للأسف."
        })

    workbook = load_workbook(file_path)
    sheet = workbook["Attendance"]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        excel_student_id = str(row[0]).strip()
        excel_date = str(row[2]).strip()

        if excel_student_id == student_id and excel_date == today:
            workbook.close()
            return jsonify({
                "reply": "تمام، حضورك اتسجل النهارده ✅"
            })

    workbook.close()

    return jsonify({
        "reply": "لا، حضورك مش متسجل للأسف."
    })


@app.route("/video_feed")
def video_feed():
    return Response(
        generate_frames(),
        mimetype="multipart/x-mixed-replace; boundary=frame"
    )


if __name__ == "__main__":
    print("🚀 Face API started on http://127.0.0.1:5000")
    app.run(host="127.0.0.1", port=5000, debug=True)